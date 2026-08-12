"""Servicios de alto nivel: calificar, persistir y emitir certificados."""
import logging
import re
from decimal import Decimal
from types import SimpleNamespace

from django.core.exceptions import ValidationError
from django.utils import timezone
from django.utils.dateparse import parse_date

from apps.origin import engine
from apps.origin.models import Certificate, OriginAnalysis, Qualification

# Datos mínimos del certificador (parte del Anexo 5-A del T-MEC).
REQUIRED_PARTY_FIELDS = ["nombre", "direccion", "pais", "email", "telefono"]


def calculate_bom_origin(bom):
    """Calcula el origen del producto terminado a partir del BOM que subió el
    proveedor, aplicando la regla (PSR) de la fracción del producto y tratado.

    Un componente se considera originario si su país está entre los miembros del
    tratado. Soporta CTC (salto arancelario), VCR y combinaciones. Guarda el
    resultado y la traza en el propio BOM y lo devuelve."""
    sr = bom.solicitation
    treaty = sr.treaty
    product = sr.product
    members = treaty.member_countries or []

    rule = bom.rule or engine.find_rule(treaty, product.hs_code or "")
    if rule is None:
        return _save_bom_result(bom, None, {
            "status": "INSUFFICIENT", "criterion": "", "rvc_value": None,
            "detail": {"error": f"No hay regla de origen para HS "
                                f"{product.hs_code} en {treaty.code}."}})

    total = Decimal("0")
    vnm = Decimal("0")
    lines = []
    for l in bom.lines.all():
        val = (l.unit_price or Decimal("0")) * (l.quantity or Decimal("0"))
        originating = bool(l.country and l.country.upper() in members)
        total += val
        if not originating:
            vnm += val
        lines.append({
            "sku": l.part_number, "hs_code": l.hs_code or "",
            "quantity": str(l.quantity), "unit_cost": str(l.unit_price),
            "line_value": str(val), "originating": originating,
            "country": l.country, "has_evidence": l.has_origin_evidence,
        })

    transaction_value = product.unit_cost if product.unit_cost else total
    # Base del VCR según el método elegido por el proveedor.
    method = bom.rvc_method or "transaction"
    if method == "net_cost" and bom.net_cost and Decimal(bom.net_cost) > 0:
        rvc_base = Decimal(bom.net_cost)
    else:
        rvc_base = transaction_value
        if method == "net_cost":
            method = "transaction"  # sin costo neto capturado, se usa transacción
    fake_product = SimpleNamespace(hs_code=product.hs_code or "")
    rt = rule.rule_type
    params = dict(rule.params or {})
    params["rvc_method"] = method  # el método lo decide el proveedor
    shift_level = params.get("shift_level", "CTH")
    de_minimis = params.get("de_minimis", treaty.de_minimis_pct)
    if (params.get("extra") or {}).get("de_minimis_excluded"):
        de_minimis = 0  # la PSR excluye la tolerancia (p.ej. cap. 1-5 agrícolas)
    except_codes = params.get("ctc_except", [])
    detail = {"rule": str(rule), "rule_type": rt, "bom": lines,
              "total_value": str(total)}

    ctc_pass = rvc_pass = None
    rvc_value = None
    if rt in ("CTC", "CTC_OR_RVC", "CTC_AND_RVC"):
        ctc_pass, ctc_detail = engine._check_tariff_shift(
            fake_product, lines, shift_level, de_minimis, total, except_codes=except_codes)
        detail["tariff_shift"] = ctc_detail
    if rt in ("RVC", "CTC_OR_RVC", "CTC_AND_RVC"):
        rvc_pass, rvc_value, rvc_detail = engine._check_rvc(
            treaty, params, rvc_base, vnm)
        detail["rvc"] = rvc_detail

    note = engine.automotive_note(params, product.hs_code or "", treaty)
    if note:
        detail["automotive_regime"] = note

    if rt == "WO":
        passed, criterion = (bool(lines) and all(ln["originating"] for ln in lines)), "WO"
    elif rt == "CTC":
        passed, criterion = bool(ctc_pass), "CTC"
    elif rt == "RVC":
        passed, criterion = bool(rvc_pass), "RVC"
    elif rt == "CTC_OR_RVC":
        passed = bool(ctc_pass) or bool(rvc_pass)
        criterion = "CTC" if ctc_pass else ("RVC" if rvc_pass else "CTC_OR_RVC")
    else:  # CTC_AND_RVC
        passed = bool(ctc_pass) and bool(rvc_pass)
        criterion = "CTC_AND_RVC"

    result = engine.apply_core_part_review(product.hs_code or "", {
        "status": "QUALIFIES" if passed else "DOES_NOT",
        "criterion": criterion, "rvc_value": rvc_value, "detail": detail}, treaty)
    return _save_bom_result(bom, rule, result)


def _save_bom_result(bom, rule, result):
    if rule:
        bom.rule = rule
    bom.origin_status = result["status"]
    bom.criterion = result["criterion"]
    bom.rvc_value = result["rvc_value"]
    bom.detail = result["detail"]
    bom.computed_at = timezone.now()
    bom.save(update_fields=["rule", "origin_status", "criterion", "rvc_value",
                            "detail", "computed_at", "updated_at"])
    return result


def _resolve_component_origin(bc, treaty, default_as_of, visited):
    """Resuelve el origen de UN insumo del BOM, según su toggle:

    - MANUAL: usa el país capturado (originario si es país miembro).
    - DECLARACIÓN: usa la declaración del proveedor (periodo o más reciente).
    - Si NO hay declaración y el insumo es un SUBENSAMBLE con su propio BOM
      (fabricado en casa), se calcula su origen RECURSIVAMENTE y, si califica,
      su valor TOTAL cuenta como originario (roll-up).
    - Si no, cae a país miembro y, en último caso, a no originario.
    """
    from apps.catalog.models import SupplierDeclaration

    comp = bc.component
    value = (comp.unit_cost or Decimal("0")) * (bc.quantity or Decimal("0"))
    members = [c.upper() for c in (treaty.member_countries or [])]

    if bc.origin_mode == "manual":
        country = (bc.manual_country or "").upper()
        if country:
            return {"originating": country in members, "country": country,
                    "value": value, "source": f"País manual: {country}"}
        return {"originating": bool(bc.manual_is_originating), "country": "",
                "value": value, "source": "Captura manual de la empresa"}

    as_of = bc.origin_as_of or default_as_of
    qs = SupplierDeclaration.objects.filter(product=comp, treaty=treaty)
    if comp.supplier_id:
        qs = qs.filter(supplier_id=comp.supplier_id)
    decl = None
    if as_of:
        decl = qs.filter(valid_from__lte=as_of, valid_to__gte=as_of).order_by("-valid_from").first()
    if decl is None:
        decl = qs.order_by("-valid_from", "-created_at").first()
    if decl:
        return {"originating": decl.is_originating,
                "country": (decl.country_of_origin or "").upper(), "value": value,
                "source": f"Declaración del proveedor {decl.valid_from} → {decl.valid_to}"}

    # Subensamble fabricado en casa: calificarlo con SU propio BOM (roll-up).
    if comp.id not in visited and comp.bom_components.exists():
        sub = _evaluate_product(comp, treaty, as_of, visited)
        orig = sub["status"] == "QUALIFIES"
        rvc = sub.get("rvc_value")
        return {"originating": orig, "country": (comp.country_of_origin or "").upper(),
                "value": value,
                "source": ("Subensamble calculado (roll-up): "
                           + ("originario" if orig else "no originario")
                           + (f" · VCR {rvc}%" if rvc is not None else ""))}

    if comp.country_of_origin and comp.country_of_origin.upper() in members:
        return {"originating": True, "country": comp.country_of_origin.upper(),
                "value": value, "source": "País miembro (sin declaración)"}
    return {"originating": False, "country": (comp.country_of_origin or "").upper(),
            "value": value, "source": "Sin declaración ni BOM propio"}


def bom_lines_for(product, treaty, as_of=None):
    """Devuelve (lines, total_materiales, vnm) del BOM de `product` para `treaty`,
    con el mismo formato de línea del motor (para snapshots/reportes, p.ej. el
    cálculo automotriz que necesita guardar el desglose del BOM)."""
    total = Decimal("0")
    vnm = Decimal("0")
    lines = []
    comps = product.bom_components.select_related("component", "component__supplier").all()
    for bc in comps:
        info = _resolve_component_origin(bc, treaty, as_of, set())
        val = info["value"]
        total += val
        if not info["originating"]:
            vnm += val
        lines.append({
            "sku": bc.component.sku, "description": bc.component.description,
            "hs_code": bc.component.hs_code or "", "quantity": str(bc.quantity),
            "unit_cost": str(bc.component.unit_cost), "line_value": str(val),
            "originating": info["originating"], "country": info["country"],
            "origin_source": info["source"],
            "supplier": (bc.component.supplier.name if bc.component.supplier_id else ""),
        })
    return lines, total, vnm


def _evaluate_product(product, treaty, as_of, visited):
    """Evalúa el origen de un producto a partir de su BOM, RECURSIVAMENTE: los
    subensambles con BOM propio se califican primero y, si originan, hacen
    roll-up (su valor total cuenta como originario). Detecta ciclos."""
    if product.id in visited:
        return {"status": "INSUFFICIENT", "criterion": "", "rvc_value": None, "rule": None,
                "detail": {"error": f"Ciclo en el BOM detectado en {product.sku}."}}
    visited = visited | {product.id}
    components = list(product.bom_components.select_related("component", "component__supplier").all())
    if not components:
        return {"status": "INSUFFICIENT", "criterion": "", "rvc_value": None, "rule": None,
                "detail": {"error": "El producto no tiene lista de materiales (BOM)."}}
    rule = engine.find_rule(treaty, product.hs_code or "")
    if rule is None:
        return {"status": "INSUFFICIENT", "criterion": "", "rvc_value": None, "rule": None,
                "detail": {"error": f"No hay regla de origen para HS "
                                    f"{product.hs_code} en {treaty.code}."}}

    total = Decimal("0")
    vnm = Decimal("0")
    lines = []
    for bc in components:
        info = _resolve_component_origin(bc, treaty, as_of, visited)
        val = info["value"]
        total += val
        if not info["originating"]:
            vnm += val
        lines.append({
            "sku": bc.component.sku, "description": bc.component.description,
            "hs_code": bc.component.hs_code or "", "quantity": str(bc.quantity),
            "unit_cost": str(bc.component.unit_cost), "line_value": str(val),
            "originating": info["originating"], "country": info["country"],
            "origin_source": info["source"],
            "supplier": (bc.component.supplier.name if bc.component.supplier_id else ""),
        })

    # Costo neto del bien = materiales (BOM) + mano de obra/costos de conversión.
    # La conversión es valor agregado REGIONAL (originario): suma a la base del VCR
    # pero NO al VNM. Es la base del cálculo de Valor de Contenido Regional.
    conversion = Decimal(str(product.conversion_cost or 0))
    net_cost = total + conversion
    rt = rule.rule_type
    params = dict(rule.params or {})
    params.setdefault("rvc_method", "transaction")
    shift_level = params.get("shift_level", "CTH")
    de_minimis = params.get("de_minimis", treaty.de_minimis_pct)
    if (params.get("extra") or {}).get("de_minimis_excluded"):
        de_minimis = 0
    except_codes = params.get("ctc_except", [])
    fake_product = SimpleNamespace(hs_code=product.hs_code or "")
    detail = {"rule": str(rule), "rule_type": rt, "bom": lines,
              "materials_total": str(total), "conversion_cost": str(conversion),
              "total_value": str(net_cost), "vnm": str(vnm),
              "psr": {"hs_pattern": rule.hs_pattern, "rule_type": rt,
                      "description": rule.description}}

    ctc_pass = rvc_pass = None
    rvc_value = None
    if rt in ("CTC", "CTC_OR_RVC", "CTC_AND_RVC"):
        ctc_pass, ctc_detail = engine._check_tariff_shift(
            fake_product, lines, shift_level, de_minimis, total, except_codes=except_codes)
        detail["tariff_shift"] = ctc_detail
    if rt in ("RVC", "CTC_OR_RVC", "CTC_AND_RVC"):
        rvc_pass, rvc_value, rvc_detail = engine._check_rvc(treaty, params, net_cost, vnm)
        detail["rvc"] = rvc_detail
    note = engine.automotive_note(params, product.hs_code or "", treaty)
    if note:
        detail["automotive_regime"] = note

    if rt == "WO":
        passed, criterion = (bool(lines) and all(l["originating"] for l in lines)), "WO"
    elif rt == "CTC":
        passed, criterion = bool(ctc_pass), "CTC"
    elif rt == "RVC":
        passed, criterion = bool(rvc_pass), "RVC"
    elif rt == "CTC_OR_RVC":
        passed = bool(ctc_pass) or bool(rvc_pass)
        criterion = "CTC" if ctc_pass else ("RVC" if rvc_pass else "CTC_OR_RVC")
    else:  # CTC_AND_RVC
        passed = bool(ctc_pass) and bool(rvc_pass)
        criterion = "CTC_AND_RVC"

    result = {"status": "QUALIFIES" if passed else "DOES_NOT", "criterion": criterion,
              "rvc_value": rvc_value, "detail": detail, "rule": rule}
    # Core part (Anexo 4-B): SOLO T-MEC; el salto/VCR del BOM no concluye y requiere
    # el régimen automotriz. En otros tratados rige la PSR normal.
    return engine.apply_core_part_review(product.hs_code or "", result, treaty)


def calculate_product_origin(product, treaty, as_of=None, user=None):
    """Calcula el origen del producto de la EMPRESA a partir de SU BOM, con
    roll-up recursivo de subensambles. Guarda la Qualification y devuelve la traza."""
    result = _evaluate_product(product, treaty, as_of, set())
    rule = result.get("rule")
    _save_qual(product, treaty, rule, result, user)
    # Snapshot en el histórico: cada corrida queda guardada con su fecha, para poder
    # comparar resultados cuando cambian precios del producto o de los insumos del BOM.
    # Auxiliar: un fallo aquí no debe tumbar el cálculo.
    try:
        save_analysis_snapshot(product, treaty, OriginAnalysis.Kind.BOM, result, user)
    except Exception:
        logging.getLogger(__name__).exception(
            "No se pudo guardar el snapshot del análisis BOM (product=%s, treaty=%s)",
            product.pk, treaty.pk)
    # El resultado se devuelve por la API: no debe contener el objeto OriginRule
    # (no es JSON-serializable). La descripción de la regla ya va en detail["rule"].
    result.pop("rule", None)
    result["rule_id"] = rule.pk if rule else None
    return result


def save_analysis_snapshot(product, treaty, kind, result, user):
    """Guarda un registro del histórico de análisis de origen (no toca la
    Qualification vigente). `result` es la traza del motor (BOM o automotriz)."""
    detail = result.get("detail") or {}

    def _dec(v):
        try:
            return Decimal(str(v))
        except Exception:
            return None
    return OriginAnalysis.objects.create(
        tenant=product.tenant, product=product, treaty=treaty, kind=kind,
        status=result.get("status") or "", criterion=result.get("criterion") or "",
        rvc_value=_dec(result.get("rvc_value")) if result.get("rvc_value") is not None else None,
        total_value=_dec(detail.get("total_value")),
        vnm=_dec(detail.get("vnm")),
        detail=detail, computed_by=user)


def _save_qual(product, treaty, rule, result, user):
    Qualification.objects.update_or_create(
        tenant=product.tenant, product=product, treaty=treaty,
        defaults={"status": result["status"], "criterion": result["criterion"],
                  "rvc_value": result["rvc_value"], "rule_id": rule.pk if rule else None,
                  "detail": result["detail"], "computed_by": user})
    return result


def qualify_and_save(product, treaty, user=None, as_of=None):
    """Ejecuta el motor y guarda/actualiza la Qualification del producto×tratado."""
    result = engine.qualify(product, treaty, as_of=as_of)
    rule_id = result["rule_id"]
    qualification, _ = Qualification.objects.update_or_create(
        tenant=product.tenant,
        product=product,
        treaty=treaty,
        defaults={
            "status": result["status"],
            "criterion": result["criterion"],
            "rvc_value": result["rvc_value"],
            "rule_id": rule_id,
            "detail": result["detail"],
            "computed_by": user,
        },
    )
    return qualification


def _next_folio(tenant):
    """Folio simple e incremental por tenant: FTA-<tenant>-<n>."""
    n = Certificate.objects.filter(tenant=tenant).count() + 1
    return f"FTA-{tenant.id}-{n:05d}"


def issue_certificate(qualification, *, certifier_type, certifier_data,
                      exporter_data=None, producer_data=None, importer_data=None,
                      blanket_from=None, blanket_to=None, folio=None, user=None):
    """Emite un certificado de origen para una calificación.

    Reglas de negocio (T-MEC):
      - Solo se emite si la calificación CALIFICA.
      - Deben estar los datos mínimos del certificador.
      - La calificación debe tener un criterio de origen.
      - El periodo (blanket) no puede exceder 12 meses.
    Lanza ValidationError si algo falla.
    """
    if qualification.status != Qualification.Status.QUALIFIES:
        raise ValidationError(
            "No se puede emitir certificado: el producto NO califica para este tratado."
        )
    if not qualification.criterion:
        raise ValidationError("La calificación no tiene criterio de origen.")

    certifier_data = certifier_data or {}
    missing = [f for f in REQUIRED_PARTY_FIELDS if not certifier_data.get(f)]
    if missing:
        raise ValidationError(f"Faltan datos del certificador: {', '.join(missing)}.")

    # Las fechas pueden llegar como texto ISO desde la API.
    if isinstance(blanket_from, str):
        blanket_from = parse_date(blanket_from)
    if isinstance(blanket_to, str):
        blanket_to = parse_date(blanket_to)

    if blanket_from and blanket_to:
        if blanket_to < blanket_from:
            raise ValidationError("El periodo del certificado es inválido (fin antes que inicio).")
        if (blanket_to - blanket_from).days > 366:
            raise ValidationError("El periodo del certificado no puede exceder 12 meses.")

    return Certificate.objects.create(
        tenant=qualification.tenant,
        qualification=qualification,
        folio=folio or _next_folio(qualification.tenant),
        certifier_type=certifier_type,
        certifier_data=certifier_data,
        exporter_data=exporter_data or {},
        producer_data=producer_data or {},
        importer_data=importer_data or {},
        blanket_from=blanket_from,
        blanket_to=blanket_to,
        issued_by=user,
    )


def certificate_elements(certificate):
    """Devuelve los 9 elementos mínimos del T-MEC (Anexo 5-A) listos para imprimir."""
    q = certificate.qualification
    prod = q.product
    return {
        "1_tipo_certificador": certificate.get_certifier_type_display(),
        "2_certificador": certificate.certifier_data,
        "3_exportador": certificate.exporter_data,
        "4_productor": certificate.producer_data,
        "5_importador": certificate.importer_data,
        "6_descripcion_y_hs": {
            "descripcion": prod.description,
            "hs_6": (prod.hs_code or "")[:6],
        },
        "7_criterio_de_origen": q.criterion,
        "8_periodo": {"desde": certificate.blanket_from, "hasta": certificate.blanket_to},
        "9_firma_y_fecha": {
            "folio": certificate.folio,
            "emitido_por": certificate.issued_by.get_username() if certificate.issued_by else None,
            "fecha": certificate.issued_at,
        },
    }


# --- Certificado de origen del PROVEEDOR por solicitud (firmado) ---

_TREATY_LABELS = {"TMEC": "USMCA"}


def _usmca_pref(criterion, status):
    """Criterio de preferencia USMCA (A–D) desde el criterio interno. Orientativo.
    Etiquetas en INGLÉS: el certificado se emite en ese idioma."""
    if status != "QUALIFIES":
        return ("—", "Origin not confirmed")
    c = (criterion or "").upper()
    if c == "WO":
        return ("A", "Wholly obtained or produced (Art. 4.2(a))")
    if "CTC" in c or "RVC" in c or "AUTOMOTRIZ" in c:
        return ("B", "Meets the product-specific rule of origin (Annex 4-B)")
    return ("B", criterion or "Meets the applicable PSR")


# El certificado no lleva hoja de instrucciones, así que el criterio debe decir
# QUÉ regla específica aplica y CÓMO se cumplió. Todo en inglés (idioma del doc).
_SHIFT_EN = {
    "CC": "a change of chapter (CC)",
    "CTH": "a change from any other heading (CTH)",
    "CTSH": "a change from any other subheading (CTSH)",
}
_RVC_METHOD_EN = {"transaction": "transaction value", "net_cost": "net cost",
                  "build_down": "build-down", "build_up": "build-up"}


def _fmt_pct(v):
    try:
        return f"{float(v):g}"
    except (TypeError, ValueError):
        return str(v)


def _rvc_req_en(params):
    """Requisito de VCR de la regla, en inglés (soporta rvc_options múltiples)."""
    opts = params.get("rvc_options") or []
    if opts:
        return " or ".join(
            f"RVC ≥ {_fmt_pct(o.get('threshold'))}% "
            f"({_RVC_METHOD_EN.get(o.get('method'), o.get('method') or 'transaction value')})"
            for o in opts)
    thr = params.get("rvc_threshold")
    method = _RVC_METHOD_EN.get(params.get("rvc_method"), params.get("rvc_method") or "transaction value")
    if thr is None:
        return f"RVC per the treaty's general threshold ({method})"
    return f"RVC ≥ {_fmt_pct(thr)}% ({method})"


def usmca_rule_text(q):
    """Texto en inglés de la regla específica (PSR) aplicada y cómo se cumplió.
    Se imprime bajo el criterio de preferencia del certificado."""
    rule = q.rule
    if not rule or q.status != "QUALIFIES":
        return ""
    params = rule.params or {}
    # La regla puede cubrir por prefijo (hs_pattern) o por rango (hs_from–hs_to).
    hs = rule.hs_pattern or (f"{rule.hs_from}-{rule.hs_to}" if rule.hs_from else
                             (q.product.hs_code or "")[:6])
    hs_fmt = (hs[:4] + "." + hs[4:6]) if len(hs) >= 6 and hs.isdigit() else hs
    ctc = _SHIFT_EN.get(params.get("shift_level", "CTH"), params.get("shift_level") or "CTH")
    reqs = {
        "WO": "the good must be wholly obtained or produced",
        "CTC": ctc,
        "RVC": _rvc_req_en(params),
        "CTC_OR_RVC": f"{ctc}; or {_rvc_req_en(params)}",
        "CTC_AND_RVC": f"{ctc} and {_rvc_req_en(params)}",
    }
    req = reqs.get(rule.rule_type, rule.rule_type)
    rvc_txt = f"RVC {q.rvc_value}%" if q.rvc_value is not None else "RVC"
    met = {
        "WO": "wholly obtained",
        "CTC": "met by tariff classification change",
        "RVC": f"met with {rvc_txt}",
        "CTC_AND_RVC": f"met by tariff classification change and {rvc_txt}",
    }.get((q.criterion or "").upper(), "")
    txt = f"PSR (Annex 4-B) for HS {hs_fmt}: {req}" if hs_fmt else f"PSR (Annex 4-B): {req}"
    return f"{txt} — {met}" if met else txt


def certificate_importer_data(cert):
    """importer_data del certificado completado EN VIVO desde Catálogos → Clientes
    (match por RFC o nombre): si la dirección/tel/email del cliente se capturan
    DESPUÉS de emitir, los documentos ya emitidos también los imprimen."""
    im = dict(cert.importer_data or {})
    missing = [k for k in ("direccion", "telefono", "email", "pais") if not im.get(k)]
    if not missing:
        return im
    from apps.catalog.models import Party
    p = None
    if im.get("rfc"):
        p = Party.objects.filter(tenant_id=cert.tenant_id, kind="customer",
                                 tax_id=im["rfc"]).first()
    if p is None and im.get("nombre"):
        p = Party.objects.filter(tenant_id=cert.tenant_id, kind="customer",
                                 name=im["nombre"]).first()
    if p:
        vals = {"direccion": p.address, "telefono": p.phone,
                "email": p.email, "pais": p.country}
        for k in missing:
            if vals.get(k):
                im[k] = vals[k]
    return im


def usmca_method_of_qualification(q):
    """9. Method of Qualification del layout de brokers USMCA: cómo calificó el
    bien — TS (salto arancelario), NC (VCR costo neto), TV (VCR valor de
    transacción), WO (totalmente obtenido). Orientativo."""
    if q.status != "QUALIFIES":
        return ""
    c = (q.criterion or "").upper()
    if c == "WO":
        return "WO"
    if "AUTOMOTRIZ" in c:
        return "NC"  # régimen automotriz: VCR por costo neto
    params = (q.rule.params if q.rule_id else {}) or {}
    rvc = "NC" if params.get("rvc_method") == "net_cost" else "TV"
    if c == "CTC":
        return "TS"
    if c == "RVC":
        return rvc
    if c == "CTC_AND_RVC":
        return f"TS + {rvc}"
    return "TS"


# Texto COMPLETO de certificación USMCA (3 compromisos + número de páginas),
# como lo exigen los agentes aduanales; sustituye a la versión corta.
def usmca_certification_text(pages=1):
    return (
        "I CERTIFY THAT: (1) the goods described in this document qualify as originating "
        "under the United States–Mexico–Canada Agreement and the information contained in "
        "this document is true and accurate, and I assume responsibility for proving such "
        "representations and agree to maintain and present upon request or to make available "
        "during a verification visit, documentation necessary to support this certification; "
        "(2) I agree to inform, in writing, all persons to whom the certificate was given of "
        "any changes that could affect the accuracy or validity of this certificate; and "
        "(3) there has been no further production or any other operation outside the "
        "territories of the Parties other than unloading, reloading, or any other operation "
        "necessary to preserve the goods in good condition or to transport them into the "
        "territory of the importing Party, and the goods did not leave the custody of the "
        "customs authorities while outside the territories of the Parties. "
        f"THIS CERTIFICATE CONSISTS OF {pages} PAGE(S), INCLUDING ALL ATTACHMENTS.")


_MX_RFC_RE = re.compile(r"^[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3}$")


def certificate_country_of_origin(cert):
    """País de origen del bien = país del PRODUCTOR, con fallbacks para datos
    incompletos: producer.pais → certifier.pais → país del perfil de la empresa
    (vivo, por si lo capturan después de emitir) → inferido de un RFC mexicano."""
    pr = cert.producer_data or {}
    ce = cert.certifier_data or {}
    pais = (pr.get("pais") or ce.get("pais") or "").strip()
    if not pais:
        prof = getattr(cert.tenant, "profile", None)
        pais = ((getattr(prof, "country", "") or "") if prof else "").strip()
    if not pais:
        rfc = (pr.get("rfc") or ce.get("rfc") or "").strip().upper()
        if _MX_RFC_RE.match(rfc):
            pais = "MX"  # formato de RFC del SAT → productor mexicano
    return pais.upper() if len(pais) <= 3 else pais


def build_certificate_xlsx(cert):
    """Genera el Certificado de Origen en formato oficial USMCA como .xlsx (bytes).
    Mismos 7 campos que el PDF: exportador, periodo, productor, importador,
    mercancía, certificación y firma."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    q = cert.qualification
    # Documento MULTI-LÍNEA: todas las partes (certificados viejos: solo la FK).
    quals = (list(cert.qualifications.select_related("product", "treaty", "rule").all())
             or [q])
    ce = cert.certifier_data or {}
    im = certificate_importer_data(cert)
    pr = cert.producer_data or ce
    label = _TREATY_LABELS.get(q.treaty.code, q.treaty.code)
    periodo = (f"{cert.blanket_from} → {cert.blanket_to}"
               if cert.blanket_from and cert.blanket_to else "Single shipment")
    # No originario → AFFIDAVIT (Value of Originating Material / VOM). La emisión
    # garantiza estados uniformes entre las partes del documento.
    is_affidavit = q.status != "QUALIFIES"

    def _vom_for(qq):
        an = OriginAnalysis.objects.filter(
            tenant_id=cert.tenant_id, product_id=qq.product_id,
            treaty_id=qq.treaty_id).order_by("-created_at").first()
        t = an.total_value if (an and an.total_value is not None) else None
        v = (an.vnm if an and an.vnm is not None else Decimal("0"))
        return t, v
    doc_title = "AFFIDAVIT OF ORIGIN (VOM)" if is_affidavit else "CERTIFICATE OF ORIGIN"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Certificate of Origin"
    navy = PatternFill("solid", fgColor="043A70")
    grey = PatternFill("solid", fgColor="EEF2F6")
    thin = Side(style="thin", color="9CA3AF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    white = Font(bold=True, color="FFFFFF")
    for col, w in zip("ABCDEFG", (20, 26, 12, 22, 15, 15, 13)):
        ws.column_dimensions[col].width = w

    def merge(rng, value, *, fill=None, font=None, wrap=False, align=None):
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = value
        if fill:
            c.fill = fill
        if font:
            c.font = font
        c.alignment = Alignment(wrap_text=wrap, vertical="top",
                                horizontal=align or "left")
        for row in ws[rng]:
            for cc in row:
                cc.border = box

    def party_lines(d):
        return (f"Name: {d.get('nombre','—')}\nAddress: {d.get('direccion','—')}"
                f"{(' ['+d['pais']+']') if d.get('pais') else ''}\n"
                f"Tax ID: {d.get('rfc','—')}\nTel: {d.get('telefono','—')}   "
                f"E-mail: {d.get('email','—')}")

    r = 1
    merge(f"A{r}:C{r}", doc_title, font=Font(bold=True, size=14, color="043A70"))
    merge(f"D{r}:G{r}", f"Document No.: {cert.folio}", font=bold, align="right")
    r += 1
    merge(f"A{r}:C{r}", f"{label} (USMCA / T-MEC)", font=Font(italic=True, color="6B7280"))
    merge(f"D{r}:G{r}", f"Issued: {str(cert.issued_at)[:10]}", align="right")
    r += 1
    # Anexo 5-A elemento 1: tipo de certificador
    ct = cert.certifier_type
    ck = lambda t: "☑" if ct == t else "☐"
    merge(f"A{r}:G{r}", f"1. Certifier is the:   {ck('exporter')} Exporter    "
          f"{ck('producer')} Producer    {ck('importer')} Importer", font=bold)
    r += 1
    # 2 certifier / 3 exportador
    ws.row_dimensions[r].height = 76
    certifier = (f"2. Certifier\nName: {ce.get('nombre','—')}\n"
                 f"Certifier/Title: {ce.get('firmante','—')}"
                 f"{(', '+ce['cargo']) if ce.get('cargo') else ''}\n"
                 f"Address: {ce.get('direccion','—')}{(' ['+ce['pais']+']') if ce.get('pais') else ''}\n"
                 f"Tel: {ce.get('telefono','—')}   E-mail: {ce.get('email','—')}")
    merge(f"A{r}:C{r}", certifier, wrap=True); ws[f"A{r}"].font = Font(size=10)
    merge(f"D{r}:G{r}", "3. Exporter\n" + party_lines(ce), wrap=True)
    r += 1
    # 4 productor / 5 importador
    ws.row_dimensions[r].height = 76
    merge(f"A{r}:C{r}", "4. Producer\n" + (party_lines(pr) if pr is not ce else "Same as certifier"), wrap=True)
    merge(f"D{r}:G{r}", "5. Importer\n" + party_lines(im), wrap=True)
    r += 1
    # 8 blanket period + factura
    merge(f"A{r}:G{r}", f"Blanket Period: {periodo}    •    "
          f"Invoice No. (single shipment): {cert.invoice_number or '—'}", font=bold)
    r += 1
    # 6/7 mercancía
    merge(f"A{r}:G{r}", "6. Description & HS Classification  ·  7–9. Origin Criteria", fill=navy, font=white)
    r += 1
    heads = ["Serial / Part No.", "Description of Good(s)", "HS No. (6-digit)",
             "7. Origin Criterion", "8. Certification Indicator",
             "9. Method of Qualification", "Country of Origin"]
    for i, h in enumerate(heads):
        c = ws.cell(row=r, column=1 + i, value=h)
        c.fill = navy; c.font = white; c.border = box
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
    # País de origen = país del PRODUCTOR (con fallbacks), no del importador.
    pais_origen = certificate_country_of_origin(cert) or "—"
    for qq in quals:
        pp = qq.product
        hs = pp.hs_code or ""
        hs_fmt = (hs[:4] + "." + hs[4:6]) if len(hs) >= 6 else hs
        if is_affidavit:
            crit, cert_ind, method = "NOT ORIGINATING", "—", "—"
        else:
            letter_i, plabel_i = _usmca_pref(qq.criterion, qq.status)
            rule_txt = usmca_rule_text(qq)
            crit = f"{letter_i} — {plabel_i}" + (f"\n{rule_txt}" if rule_txt else
                                                 (f" · RVC {qq.rvc_value}%" if qq.rvc_value else ""))
            # 8. Certification Indicator: el certificador ¿es el productor del bien?
            cert_ind = "YES" if cert.certifier_type == "producer" else "NO"
            # 9. Method of Qualification: TS / NC / TV / WO.
            method = usmca_method_of_qualification(qq) or "—"
        ws.row_dimensions[r].height = 58
        for i, v in enumerate([pp.sku, pp.description, hs_fmt, crit, cert_ind,
                               method, pais_origen]):
            c = ws.cell(row=r, column=1 + i, value=v)
            c.border = box; c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1

    def _money(v):
        return "—" if v is None else f"${float(v):,.2f}"

    if is_affidavit:
        merge(f"A{r}:G{r}", "6. Value of Originating Material (VOM)", fill=navy, font=white)
        r += 1
        # Una fila por parte + fila TOTAL (los valores salen del último cálculo).
        for h, rng in (("Part No.", f"A{r}:B{r}"), ("Total value (net cost)", f"C{r}:D{r}"),
                       ("Non-originating (VNM)", f"E{r}:F{r}"), ("Originating (VOM)", f"G{r}:G{r}")):
            merge(rng, h, fill=grey, font=bold)
        r += 1
        tot_sum, vnm_sum, con_datos = Decimal("0"), Decimal("0"), False
        for qq in quals:
            t, v = _vom_for(qq)
            vom_i = (t - v) if t is not None else None
            if t is not None:
                con_datos = True
                tot_sum += t; vnm_sum += v
            merge(f"A{r}:B{r}", qq.product.sku)
            for rng, val in ((f"C{r}:D{r}", t), (f"E{r}:F{r}", v if t is not None else None),
                             (f"G{r}:G{r}", vom_i)):
                merge(rng, _money(val), align="right")
            r += 1
        merge(f"A{r}:B{r}", "TOTAL", font=bold)
        for rng, val in ((f"C{r}:D{r}", tot_sum if con_datos else None),
                         (f"E{r}:F{r}", vnm_sum if con_datos else None),
                         (f"G{r}:G{r}", (tot_sum - vnm_sum) if con_datos else None)):
            merge(rng, _money(val), font=bold, align="right")
        r += 1
    if is_affidavit:
        cert_txt = ("Certification. I certify that the good described does NOT qualify as originating "
                    f"under the {label}, and that the Value of Originating Material (VOM) stated is true and accurate. "
                    "This affidavit lets the recipient account for the originating content in its own RVC "
                    "determination. I assume responsibility for proving these representations.")
    else:
        # Texto COMPLETO exigido por los agentes aduanales (3 compromisos + páginas).
        cert_txt = "Certification. " + usmca_certification_text(pages=1)
    merge(f"A{r}:G{r}", cert_txt, wrap=True)
    ws.row_dimensions[r].height = 60 if is_affidavit else 118
    r += 1
    ws.row_dimensions[r].height = 60
    merge(f"A{r}:C{r}", "Authorized Signature & Date\n\n_______________________________", wrap=True)
    merge(f"D{r}:G{r}",
          f"Name & Title: {ce.get('firmante','—')}"
          f"{(', '+ce['cargo']) if ce.get('cargo') else ''}\n"
          f"Company: {ce.get('nombre','—')}\nDate: {str(cert.issued_at)[:10]}\n"
          f"Tel: {ce.get('telefono','—')}   E-mail: {ce.get('email','—')}", wrap=True)
    r += 2
    merge(f"A{r}:G{r}", f"Folio {cert.folio} · {label}. Generado por LogiQ Aduanas | FTA. "
          "Orientativo; validar con un especialista en reglas de origen.",
          font=Font(size=8, color="6B7280"), wrap=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cert_party_block(prof, fallback_name="", fallback_country=""):
    """Bloque de datos de una parte (productor/importador) desde su perfil."""
    if not prof:
        return {"nombre": fallback_name, "rfc": "", "direccion": "", "pais": fallback_country,
                "email": "", "telefono": "", "firmante": "", "cargo": ""}
    direccion = ", ".join(p for p in [prof.address, prof.city, prof.state, prof.postal_code] if p)
    return {"nombre": prof.legal_name or fallback_name, "rfc": prof.tax_id,
            "direccion": direccion, "pais": prof.country or fallback_country,
            "email": prof.contact_email, "telefono": prof.contact_phone,
            "firmante": prof.signatory_name, "cargo": prof.signatory_title}


def build_solicitation_cert_data(sr):
    """Snapshot de los datos del certificado a partir de la solicitud aceptada.
    El origen viene de la declaración manual o, si es por BOM, del cálculo del BOM."""
    from apps.catalog.models import SolicitationBOM
    supplier = sr.supplier
    sup_prof = getattr(supplier, "profile", None)
    comp_prof = getattr(sr.tenant, "profile", None)

    origin = {"is_originating": None, "country": "", "criterion": "", "rule": ""}
    decl = sr.declaration
    if decl:
        origin = {
            "is_originating": decl.is_originating,
            "country": decl.country_of_origin or "",
            "criterion": (decl.rule.rule_type if decl.rule_id else ""),
            "rule": (decl.rule.description if decl.rule_id else ""),
        }
    else:
        bom = SolicitationBOM.objects.filter(solicitation=sr).select_related("rule").first()
        if bom:
            origin = {
                "is_originating": bom.origin_status == "QUALIFIES",
                "country": "",
                "criterion": bom.criterion or bom.origin_status or "",
                "rule": (bom.rule.description if bom.rule_id else ""),
            }

    return {
        "producer": _cert_party_block(sup_prof, supplier.name, supplier.country),
        "importer": _cert_party_block(comp_prof, sr.tenant.name),
        "product": {"sku": sr.product.sku, "description": sr.product.description,
                    "hs": sr.product.hs_code or ""},
        "treaty": {"code": sr.treaty.code, "label": _TREATY_LABELS.get(sr.treaty.code, sr.treaty.code)},
        "origin": origin,
        "period": {"from": sr.period_from.isoformat() if sr.period_from else None,
                   "to": sr.period_to.isoformat() if sr.period_to else None},
    }


def _next_cert_folio(tenant):
    from apps.origin.models import SolicitationCertificate
    n = SolicitationCertificate.objects.filter(tenant=tenant).count() + 1
    return f"FTA-DO-{tenant.id}-{n:05d}"


def ensure_solicitation_certificate(sr, user, sign_method):
    """Crea (o actualiza, si aún no se firma) el certificado de la solicitud con el
    método de firma EXIGIDO por la empresa. Idempotente."""
    import secrets
    from apps.origin.models import SolicitationCertificate
    cert, created = SolicitationCertificate.objects.get_or_create(
        solicitation=sr,
        defaults={"tenant": sr.tenant, "folio": _next_cert_folio(sr.tenant),
                  "verify_token": secrets.token_urlsafe(16)[:32],
                  "data": build_solicitation_cert_data(sr),
                  "sign_method": sign_method, "requested_by": user})
    if not created and not cert.signed:
        cert.sign_method = sign_method
        cert.requested_by = user
        cert.data = build_solicitation_cert_data(sr)
        cert.save(update_fields=["sign_method", "requested_by", "data", "updated_at"])
    return cert
