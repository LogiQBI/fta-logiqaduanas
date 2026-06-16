"""Carga masiva por Excel (.xlsx): plantillas e importadores por entidad.

Cada "tipo" define sus columnas (clave interna + etiqueta + ejemplo) y una
función importadora que hace upsert y devuelve un resumen {creados, actualizados,
errores}. Se usa para catálogos (productos/insumos, proveedores, clientes) y BOM.
"""
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from django.db import transaction
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from apps.catalog.uom import clean_uom


def _dec(v):
    try:
        return Decimal(str(v).strip()) if v not in (None, "") else Decimal("0")
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _digits(v, n=10):
    return "".join(c for c in str(v or "") if c.isdigit())[:n]


def _iso2(v):
    return "".join(c for c in str(v or "") if c.isalpha()).upper()[:2]


# Mapeo de tipo de producto en español -> clave del modelo.
KIND_MAP = {
    "material": "material", "insumo": "material", "materia prima": "material",
    "subensamble": "subassembly", "subassembly": "subassembly",
    "terminado": "finished", "producto terminado": "finished", "finished": "finished",
}


def _has(v):
    """¿La celda trae un valor? (para no pisar datos existentes con celdas vacías)."""
    return v is not None and str(v).strip() != ""


def import_products(tenant, rows, user):
    """Carga/actualiza números de parte. Si el SKU ya existe NO se duplica: se
    actualizan los campos que el Excel trae con valor (precio, descripción, HS,
    país, proveedor…), dejando intactos los que vengan vacíos. Si no existe, se crea."""
    from apps.catalog.models import Party, Product, ProductChangeLog, log_product_changes
    res = {"creados": 0, "actualizados": 0, "omitidos": 0, "errores": [], "advertencias": [],
           "creados_skus": [], "actualizados_skus": [], "sin_precio_skus": [], "duplicados_skus": []}
    seen = set()  # SKUs ya procesados en ESTE archivo (para no cargar duplicados)
    for i, r in enumerate(rows, start=2):
        # El número de parte se normaliza a MAYÚSCULAS y el match es sin distinguir
        # mayúsc/minúsc, para que re-subir actualice aunque cambie el case.
        sku = str(r.get("sku") or "").strip().upper()
        if not sku:
            res["errores"].append({"fila": i, "error": "Falta SKU."}); continue
        # SKU repetido dentro del mismo archivo: se importa solo la PRIMERA aparición.
        if sku in seen:
            if len(res["duplicados_skus"]) < 200:
                res["duplicados_skus"].append(sku)
            res["errores"].append({"fila": i,
                "error": f"“{sku}” está repetido en el archivo; se ignoró esta fila "
                         "(se usó la primera). Deja un solo renglón por número de parte."})
            continue
        seen.add(sku)
        try:
            adv = None
            with transaction.atomic():
                # Proveedor (solo si la fila trae código): se busca o se precarga.
                supplier = None
                scode = str(r.get("proveedor_codigo") or "").strip()
                if scode:
                    supplier = Party.objects.filter(
                        tenant=tenant, kind=Party.Kind.SUPPLIER, code__iexact=scode).first()
                    if not supplier:
                        supplier = Party.objects.create(
                            tenant=tenant, kind=Party.Kind.SUPPLIER, code=scode, name=scode)
                        adv = f"Proveedor '{scode}' precargado automáticamente; complétalo en Proveedores."

                existing = Product.objects.filter(tenant=tenant, sku__iexact=sku).first()
                if existing:
                    # Snapshot para registrar el histórico de precio/origen.
                    before = {"unit_cost": existing.unit_cost, "currency": existing.currency,
                              "country_of_origin": existing.country_of_origin}
                    # Actualiza SOLO lo que venga con valor (no pisa con vacíos).
                    if _has(r.get("descripcion")):
                        existing.description = str(r.get("descripcion")).strip()
                    if _has(r.get("tipo")):
                        existing.kind = KIND_MAP.get(str(r.get("tipo")).strip().lower(), existing.kind)
                    if _has(r.get("hs_code")):
                        existing.hs_code = _digits(r.get("hs_code"), 8)
                    if _has(r.get("costo_unitario")):
                        existing.unit_cost = _dec(r.get("costo_unitario"))
                    if _has(r.get("moneda")):
                        existing.currency = str(r.get("moneda")).strip().upper()[:3]
                    if _has(r.get("pais_origen")):
                        existing.country_of_origin = _iso2(r.get("pais_origen"))
                    if supplier:
                        existing.supplier = supplier
                    existing.save()
                    log_product_changes(
                        product=existing, before=before,
                        after={"unit_cost": existing.unit_cost, "currency": existing.currency,
                               "country_of_origin": existing.country_of_origin},
                        source=ProductChangeLog.Source.BULK, user=user)
                    res["actualizados"] += 1
                    if len(res["actualizados_skus"]) < 200:
                        res["actualizados_skus"].append(existing.sku)
                    final = existing
                else:
                    kind = KIND_MAP.get(str(r.get("tipo") or "material").strip().lower(), "material")
                    final = Product.objects.create(
                        tenant=tenant, sku=sku,
                        description=str(r.get("descripcion") or "").strip(),
                        kind=kind, hs_code=_digits(r.get("hs_code"), 8),
                        unit_cost=_dec(r.get("costo_unitario")),
                        currency=(str(r.get("moneda") or "USD").strip().upper() or "USD")[:3],
                        country_of_origin=_iso2(r.get("pais_origen")), supplier=supplier)
                    res["creados"] += 1
                    if len(res["creados_skus"]) < 200:
                        res["creados_skus"].append(final.sku)
                # Marca los que quedan SIN precio (precio 0): la fila no traía costo.
                if (not final.unit_cost) and len(res["sin_precio_skus"]) < 200:
                    res["sin_precio_skus"].append(final.sku)
            if adv:
                res["advertencias"].append({"fila": i, "error": adv})
        except Exception as e:
            res["errores"].append({"fila": i, "error": f"No se pudo guardar “{sku}”: {str(e)[:120]}"})
    return res


def preview_products(tenant, rows):
    """Analiza el archivo SIN guardar: cuántos números de parte ya existen y
    cuántos son nuevos (para confirmar antes de importar)."""
    from apps.catalog.models import Product
    skus = []
    dup = []  # SKUs repetidos dentro del mismo archivo
    seen = set()
    for r in rows:
        s = str(r.get("sku") or "").strip().upper()
        if not s:
            continue
        if s in seen:
            if s not in dup:
                dup.append(s)
        else:
            seen.add(s); skus.append(s)
    # Comparación sin distinguir mayúsc/minúsc: normaliza lo guardado a MAYÚSCULAS.
    existentes = {sk.upper() for sk in Product.objects.filter(tenant=tenant)
                  .values_list("sku", flat=True)}
    ex = [s for s in skus if s in existentes]
    nuevos = [s for s in skus if s not in existentes]
    return {"total": len(skus), "existentes": len(ex), "nuevos": len(nuevos),
            "existentes_skus": ex[:40],
            "duplicados": len(dup), "duplicados_skus": dup[:40]}


def _import_parties(tenant, rows, kind):
    from apps.catalog.models import Party
    res = {"creados": 0, "actualizados": 0, "errores": []}
    for i, r in enumerate(rows, start=2):
        name = str(r.get("nombre") or "").strip()
        if not name:
            res["errores"].append({"fila": i, "error": "Falta el nombre."}); continue
        code = str(r.get("codigo") or "").strip()
        defaults = {
            "country": _iso2(r.get("pais")), "tax_id": str(r.get("rfc") or "").strip(),
            "email": str(r.get("email") or "").strip(), "phone": str(r.get("telefono") or "").strip(),
        }
        # Buscar existente por código (si hay) o por nombre, dentro del tenant y tipo.
        existing = None
        if code:
            existing = Party.objects.filter(tenant=tenant, kind=kind, code__iexact=code).first()
        if not existing:
            existing = Party.objects.filter(tenant=tenant, kind=kind, name__iexact=name).first()
        if existing:
            existing.name = name
            if code:
                existing.code = code
            for k, v in defaults.items():
                setattr(existing, k, v)
            existing.save()
            res["actualizados"] += 1
        else:
            Party.objects.create(tenant=tenant, kind=kind, name=name, code=code, **defaults)
            res["creados"] += 1
    return res


def import_suppliers(tenant, rows, user):
    from apps.catalog.models import Party
    return _import_parties(tenant, rows, Party.Kind.SUPPLIER)


def import_customers(tenant, rows, user):
    from apps.catalog.models import Party
    return _import_parties(tenant, rows, Party.Kind.CUSTOMER)


def import_bom(tenant, rows, user):
    from apps.catalog.models import BOMComponent, Product
    res = {"creados": 0, "actualizados": 0, "errores": []}
    for i, r in enumerate(rows, start=2):
        psku = str(r.get("producto_sku") or "").strip().upper()
        csku = str(r.get("insumo_sku") or "").strip().upper()
        if not psku or not csku:
            res["errores"].append({"fila": i, "error": "Falta producto_sku o insumo_sku."}); continue
        parent = Product.objects.filter(tenant=tenant, sku__iexact=psku).first()
        component = Product.objects.filter(tenant=tenant, sku__iexact=csku).first()
        if not parent:
            res["errores"].append({"fila": i, "error": f"Producto '{psku}' no existe."}); continue
        if not component:
            res["errores"].append({"fila": i, "error": f"Insumo '{csku}' no existe."}); continue
        if parent.id == component.id:
            res["errores"].append({"fila": i, "error": "El producto no puede ser su propio insumo."}); continue
        pais = _iso2(r.get("pais_origen"))
        defaults = {"quantity": _dec(r.get("cantidad")) or Decimal("1")}
        uom = clean_uom(r.get("uom"))
        if uom:
            defaults["uom"] = uom
        if pais:
            defaults["origin_mode"] = "manual"
            defaults["manual_country"] = pais
        _, created = BOMComponent.objects.update_or_create(
            tenant=tenant, parent=parent, component=component, defaults=defaults)
        res["creados" if created else "actualizados"] += 1
    return res


def import_supplier_assign(tenant, rows, user):
    """Asigna el proveedor a números de parte ya existentes, por layout
    (num_parte + código de proveedor). Precarga el proveedor si no existe."""
    from apps.catalog.models import Party, Product
    res = {"creados": 0, "actualizados": 0, "errores": [], "advertencias": []}
    for i, r in enumerate(rows, start=2):
        sku = str(r.get("num_parte") or "").strip().upper()
        scode = str(r.get("codigo_proveedor") or "").strip()
        if not sku or not scode:
            res["errores"].append({"fila": i, "error": "Falta num_parte o codigo_proveedor."}); continue
        try:
            adv = None
            with transaction.atomic():
                product = Product.objects.filter(tenant=tenant, sku__iexact=sku).first()
                if not product:
                    res["errores"].append({"fila": i, "error": f"Número de parte '{sku}' no existe."}); continue
                supplier = Party.objects.filter(
                    tenant=tenant, kind=Party.Kind.SUPPLIER, code__iexact=scode).first()
                if not supplier:
                    supplier = Party.objects.create(
                        tenant=tenant, kind=Party.Kind.SUPPLIER, code=scode, name=scode)
                    adv = f"Proveedor '{scode}' precargado automáticamente; complétalo en Proveedores."
                product.supplier = supplier
                product.save(update_fields=["supplier", "updated_at"])
            res["actualizados"] += 1
            if adv:
                res["advertencias"].append({"fila": i, "error": adv})
        except Exception as e:
            res["errores"].append({"fila": i, "error": f"No se pudo asignar “{sku}”: {str(e)[:120]}"})
    return res


# Texto de ayuda reutilizable.
_UOM_HELP = ("Unidad de medida: código de 2 caracteres del catálogo "
             "(PZ pieza, UN unidad, KG, GR, TN, MT, CM, MM, M2, M3, LT, ML, "
             "PR par, JG juego, CJ caja, RL rollo, HJ hoja…). Déjalo vacío si no aplica.")
_HS_HELP = "Fracción arancelaria; 6 a 8 dígitos (puedes escribir solo números, ej. 720839)."
_ISO2_HELP = "País en código ISO de 2 letras (MX, US, CN, KR…)."

# Definición de cada tipo de carga: hoja, columnas (clave, etiqueta, ejemplo),
# importador, instrucciones generales y ayuda por columna ({req, help}).
SPECS = {
    "supplier_assign": {
        "sheet": "Asignar proveedor",
        "columns": [
            ("num_parte", "Núm. de parte (SKU)", "MAT-001"),
            ("codigo_proveedor", "Código de proveedor", "ST01"),
        ],
        "importer": import_supplier_assign,
        "instructions": (
            "Asigna (o reasigna) el proveedor de números de parte que YA existen en tu catálogo.\n"
            "El número de parte debe existir; el proveedor se busca por su código y, si no existe, "
            "se precarga automáticamente para que lo completes después en Proveedores.\n"
            "Una fila por número de parte. La fila de ejemplo (fila 2) puedes borrarla o sobreescribirla."),
        "help": {
            "num_parte": {"req": True, "help": "SKU del número de parte tal como está en tu catálogo."},
            "codigo_proveedor": {"req": True, "help": "Código del proveedor (ej. ST01). Si no existe, se precarga."},
        },
    },
    "products": {
        "sheet": "Insumos y productos",
        "columns": [
            ("sku", "SKU / Núm. de parte", "MAT-001"),
            ("descripcion", "Descripción", "Lámina de acero"),
            ("tipo", "Tipo (material/subensamble/terminado)", "material"),
            ("hs_code", "Fracción HS (6 díg.)", "720839"),
            ("costo_unitario", "Costo unitario", "30.00"),
            ("moneda", "Moneda", "USD"),
            ("pais_origen", "País de origen (ISO-2)", "KR"),
            ("proveedor_codigo", "Código de proveedor (opcional)", "ST01"),
        ],
        "importer": import_products,
        "instructions": (
            "Da de alta o ACTUALIZA números de parte (insumos y productos).\n"
            "Si el SKU ya existe NO se duplica: se actualizan los datos que traiga el archivo "
            "(precio, descripción, HS, país…) y se respetan las celdas que dejes vacías.\n"
            "Cada cambio de precio o de país queda en el histórico del número de parte.\n"
            "Una fila por número de parte. La fila de ejemplo (fila 2) puedes borrarla o sobreescribirla."),
        "help": {
            "sku": {"req": True, "help": "Clave única del número de parte. Si ya existe, se actualiza."},
            "descripcion": {"req": False, "help": "Nombre o descripción del insumo/producto."},
            "tipo": {"req": False, "help": "material, subensamble o terminado. Si se omite, queda como material."},
            "hs_code": {"req": False, "help": _HS_HELP},
            "costo_unitario": {"req": False, "help": "Costo unitario; solo número (ej. 30.00)."},
            "moneda": {"req": False, "help": "Moneda de 3 letras (USD, MXN…). Por defecto USD."},
            "pais_origen": {"req": False, "help": _ISO2_HELP},
            "proveedor_codigo": {"req": False, "help": "Código del proveedor que surte este número de parte. Si no existe, se precarga."},
        },
    },
    "suppliers": {
        "sheet": "Proveedores",
        "columns": [
            ("nombre", "Nombre / Razón social", "Componentes MX"),
            ("codigo", "Código de proveedor", "CMX01"),
            ("pais", "País (ISO-2)", "MX"),
            ("rfc", "RFC / Tax ID", "CMX010101AB1"),
            ("email", "Email", "contacto@proveedor.com"),
            ("telefono", "Teléfono", "8181818181"),
        ],
        "importer": import_suppliers,
        "instructions": (
            "Da de alta o actualiza tu padrón de PROVEEDORES.\n"
            "Se busca por código (si lo traes) o por nombre dentro de tu empresa; si coincide, se actualiza.\n"
            "Una fila por proveedor. La fila de ejemplo (fila 2) puedes borrarla o sobreescribirla."),
        "help": {
            "nombre": {"req": True, "help": "Razón social o nombre del proveedor."},
            "codigo": {"req": False, "help": "Clave que tú le asignas (ej. CMX01). Sirve para ligar números de parte."},
            "pais": {"req": False, "help": _ISO2_HELP},
            "rfc": {"req": False, "help": "RFC o Tax ID del proveedor."},
            "email": {"req": False, "help": "Correo de contacto."},
            "telefono": {"req": False, "help": "Teléfono de contacto."},
        },
    },
    "customers": {
        "sheet": "Clientes",
        "columns": [
            ("nombre", "Nombre / Razón social", "Importadora USA Inc"),
            ("codigo", "Código (opcional)", "CLI01"),
            ("pais", "País (ISO-2)", "US"),
            ("rfc", "RFC / Tax ID", "US-99-123"),
            ("email", "Email", "compras@cliente.com"),
            ("telefono", "Teléfono", "+1 555 0100"),
        ],
        "importer": import_customers,
        "instructions": (
            "Da de alta o actualiza tu padrón de CLIENTES.\n"
            "Se busca por código (si lo traes) o por nombre dentro de tu empresa; si coincide, se actualiza.\n"
            "Una fila por cliente. La fila de ejemplo (fila 2) puedes borrarla o sobreescribirla."),
        "help": {
            "nombre": {"req": True, "help": "Razón social o nombre del cliente."},
            "codigo": {"req": False, "help": "Clave opcional que tú le asignas (ej. CLI01)."},
            "pais": {"req": False, "help": _ISO2_HELP},
            "rfc": {"req": False, "help": "RFC o Tax ID del cliente."},
            "email": {"req": False, "help": "Correo de contacto."},
            "telefono": {"req": False, "help": "Teléfono de contacto."},
        },
    },
    "bom": {
        "sheet": "Lista de materiales (BOM)",
        "columns": [
            ("producto_sku", "SKU del producto (padre)", "FG-AUTO-01"),
            ("insumo_sku", "SKU del insumo (componente)", "MAT-001"),
            ("cantidad", "Cantidad", "1"),
            ("uom", "Unidad de medida (PZ, KG, MT…)", "PZ"),
            ("pais_origen", "País de origen manual (opcional, ISO-2)", "MX"),
        ],
        "importer": import_bom,
        "instructions": (
            "Arma las listas de materiales (BOM): cada fila liga un PRODUCTO (padre) con uno de "
            "sus INSUMOS (componente), ambos por SKU. Ambos deben existir ya en tu catálogo.\n"
            "Repite el SKU del producto en varias filas para agregarle varios insumos.\n"
            "El país de origen es opcional (deja el componente en modo manual con ese país).\n"
            "La fila de ejemplo (fila 2) puedes borrarla o sobreescribirla."),
        "help": {
            "producto_sku": {"req": True, "help": "SKU del producto terminado o subensamble (el padre del BOM)."},
            "insumo_sku": {"req": True, "help": "SKU del insumo/componente que lo integra."},
            "cantidad": {"req": False, "help": "Cantidad del insumo por unidad del producto (ej. 1, 2.5). Por defecto 1."},
            "uom": {"req": False, "help": _UOM_HELP},
            "pais_origen": {"req": False, "help": "Opcional. " + _ISO2_HELP + " Fija el origen manual del componente."},
        },
    },
}

# Columnas para responder una solicitud por BOM (proveedor).
BOM_RESPONSE_COLUMNS = [
    ("num_parte", "Número de parte", "CMP-001"),
    ("descripcion", "Descripción", "Resistencia"),
    ("hs_code", "Fracción HS (6 díg.)", "853321"),
    ("precio_unitario", "Precio unitario", "1.50"),
    ("cantidad", "Cantidad", "10"),
    ("uom", "Unidad de medida (PZ, KG, MT…)", "PZ"),
    ("pais", "País de origen (ISO-2)", "CN"),
    ("evidencia", "¿Tiene evidencia? (si/no)", "no"),
]

BOM_RESPONSE_INSTRUCTIONS = (
    "Responde la solicitud de tu cliente subiendo la lista de materiales (BOM) del número de parte.\n"
    "Una fila por componente. Indica para cada uno su fracción HS, precio, cantidad, unidad de medida, "
    "país de origen y si cuentas con documento/evidencia de origen.\n"
    "Al importar se reemplazan las líneas actuales del BOM; después revisas, calculas el origen y envías.\n"
    "La fila de ejemplo (fila 2) puedes borrarla o sobreescribirla.")

BOM_RESPONSE_HELP = {
    "num_parte": {"req": True, "help": "Número de parte del componente."},
    "descripcion": {"req": False, "help": "Descripción del componente."},
    "hs_code": {"req": False, "help": _HS_HELP + " Necesaria para evaluar el salto arancelario (CTH)."},
    "precio_unitario": {"req": False, "help": "Precio unitario del componente; solo número (ej. 1.50)."},
    "cantidad": {"req": False, "help": "Cantidad utilizada del componente; solo número (ej. 10)."},
    "uom": {"req": False, "help": _UOM_HELP},
    "pais": {"req": False, "help": _ISO2_HELP},
    "evidencia": {"req": False, "help": "¿Tienes documento/certificado que respalde el origen? Escribe sí o no."},
}


def make_template(columns, sheet_name, instructions="", col_help=None):
    """Crea un .xlsx con: (1) una hoja "Instrucciones" de llenado y (2) la hoja de
    datos con encabezados resaltados + fila de ejemplo. Cada encabezado lleva un
    comentario con su ayuda. Devuelve bytes.

    `instructions`: texto general (líneas separadas por \\n).
    `col_help`: dict {clave: {"req": bool, "help": str}}."""
    col_help = col_help or {}
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="043A70")

    # --- Hoja de datos (la que se llena e importa) ---
    ws = wb.active
    ws.title = sheet_name[:31]
    for col, (key, label, example) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.fill = fill
        ws.cell(row=2, column=col, value=example)
        ws.column_dimensions[c.column_letter].width = max(14, len(label) + 2)
        info = col_help.get(key)
        if info:
            txt = ("OBLIGATORIO. " if info.get("req") else "Opcional. ") + info.get("help", "")
            cm = Comment(txt, "FTA")
            cm.width, cm.height = 300, 130
            c.comment = cm

    # --- Hoja de instrucciones (primera y visible al abrir) ---
    ins = wb.create_sheet("Instrucciones", 0)
    ins["A1"] = "LogiQ Aduanas | FTA — Instrucciones de llenado"
    ins["A1"].font = Font(bold=True, size=14, color="043A70")
    ins["A2"] = f"Plantilla: {sheet_name}. Captura los datos en la pestaña «{sheet_name[:31]}»."
    ins["A2"].font = Font(italic=True, color="6B7280")
    r = 4
    for para in (instructions or "").split("\n"):
        cell = ins.cell(row=r, column=1, value=para)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1
    for j, h in enumerate(["Columna", "¿Obligatoria?", "Cómo llenarla", "Ejemplo"], start=1):
        c = ins.cell(row=r, column=j, value=h)
        c.font = header_font
        c.fill = fill
    r += 1
    for key, label, example in columns:
        info = col_help.get(key, {})
        ins.cell(row=r, column=1, value=label)
        ins.cell(row=r, column=2, value="Sí" if info.get("req") else "No")
        hc = ins.cell(row=r, column=3, value=info.get("help", ""))
        hc.alignment = Alignment(wrap_text=True, vertical="top")
        ins.cell(row=r, column=4, value=str(example))
        r += 1
    ins.column_dimensions["A"].width = 36
    ins.column_dimensions["B"].width = 13
    ins.column_dimensions["C"].width = 66
    ins.column_dimensions["D"].width = 18
    wb.active = 0  # abrir mostrando las instrucciones

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_rows(file, columns):
    """Lee el .xlsx subido. Busca —en cualquier hoja— la fila de ENCABEZADOS que
    coincide con las ETIQUETAS de la plantilla (ignora la hoja de Instrucciones) y
    devuelve los renglones de datos como dicts con las CLAVES internas."""
    wb = openpyxl.load_workbook(file, data_only=True)
    label_to_key = {label.strip().lower(): key for key, label, _ in columns}

    found = None  # (worksheet, índice_fila_encabezado, idx_to_key)
    for ws in wb.worksheets:
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if not row:
                continue
            idx_to_key = {}
            for idx, h in enumerate(row):
                if h is None:
                    continue
                key = label_to_key.get(str(h).strip().lower())
                if key is not None:
                    idx_to_key[idx] = key
            # La fila de encabezados tiene varias etiquetas conocidas a la vez.
            if len(idx_to_key) >= min(2, len(columns)):
                found = (ws, r_idx, idx_to_key)
                break
        if found:
            break
    if not found:
        return []

    ws, header_idx, idx_to_key = found
    out = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= header_idx:
            continue
        if row is None or all(v in (None, "") for v in row):
            continue
        d = {idx_to_key[idx]: (row[idx] if idx < len(row) else None) for idx in idx_to_key}
        out.append(d)
    return out
