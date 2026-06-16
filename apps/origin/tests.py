"""Tests de regresión de los arreglos del motor de origen y la carga masiva.

Cubren bugs concretos resueltos: el 500 por OriginRule no serializable, el
régimen automotriz (core parts → AUTO_REVIEW), el histórico de precio sin falsos
positivos por precisión, y la normalización/match de SKU en la carga masiva.
"""
from decimal import Decimal

from django.test import TestCase
from rest_framework.renderers import JSONRenderer

from apps.catalog.models import (
    BOMComponent, Product, ProductChangeLog, log_product_changes,
)
from apps.catalog.uom import clean_uom
from apps.origin import engine
from apps.origin.bulk import import_products
from apps.origin.services import calculate_product_origin
from apps.tenants.models import Tenant
from apps.treaties.models import OriginRule, Treaty

from datetime import date

from django.contrib.auth import get_user_model
from rest_framework.test import APIRequestFactory, force_authenticate

from apps.catalog.models import (
    Party, SupplierDeclaration, SolicitationRequest, SupplierProfile,
)
from apps.origin.models import SolicitationCertificate
from apps.origin.views import SolicitationCertificateViewSet, SolicitationRequestViewSet
from apps.tenants.models import Membership


class PureHelpersTest(TestCase):
    def test_core_part_code(self):
        self.assertEqual(engine.core_part_code("870880"), "870880")   # suspensión
        self.assertEqual(engine.core_part_code("8708.80"), "870880")  # con punto
        self.assertEqual(engine.core_part_code("870710"), "8707")     # carrocería (partida)
        self.assertIsNone(engine.core_part_code("830230"))            # no automotriz core
        self.assertIsNone(engine.core_part_code(""))

    def test_clean_uom(self):
        self.assertEqual(clean_uom("kg"), "KG")
        self.assertEqual(clean_uom(" Pz "), "PZ")
        self.assertEqual(clean_uom("ZZ"), "")   # fuera del catálogo
        self.assertEqual(clean_uom(None), "")


class OriginEngineTest(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="ACME", slug="acme")
        self.treaty = Treaty.objects.create(
            code="TMEC", name="USMCA", member_countries=["MX", "US", "CA"])
        OriginRule.objects.create(treaty=self.treaty, hs_pattern="870880",
                                  rule_type="CTC", params={"shift_level": "CTH"})
        OriginRule.objects.create(treaty=self.treaty, hs_pattern="940360",
                                  rule_type="CTC", params={"shift_level": "CTH"})

    def _product_with_bom(self, hs):
        p = Product.objects.create(tenant=self.tenant, sku="FG", description="x",
                                   kind="finished", hs_code=hs, unit_cost=Decimal("24.26"))
        for i, (chs, cost, q, country) in enumerate([
            ("401699", "0", 2, "PL"), ("830230", "2.0040", 2, "US"),
            ("730459", "1.3880", 1, "KR")]):
            comp = Product.objects.create(tenant=self.tenant, sku=f"C{i}", description="c",
                                          kind="material", hs_code=chs, unit_cost=Decimal(cost))
            BOMComponent.objects.create(tenant=self.tenant, parent=p, component=comp,
                                        quantity=Decimal(q), origin_mode="manual",
                                        manual_country=country)
        return p

    def test_core_part_marks_auto_review_and_is_json_serializable(self):
        """8708.80 (suspensión) NO debe 'Calificar' por CTH: AUTO_REVIEW.
        Y el resultado debe poder serializarse a JSON (regresión del 500)."""
        p = self._product_with_bom("870880")
        result = calculate_product_origin(p, self.treaty, user=None)
        self.assertEqual(result["status"], "AUTO_REVIEW")
        self.assertIn("automotive_core", result["detail"])
        # El aviso de core part reemplaza al genérico de cap. 87 (no se duplican).
        self.assertNotIn("automotive_regime", result["detail"])
        self.assertNotIn("rule", result)  # el objeto OriginRule no debe estar en el payload
        JSONRenderer().render(result)     # no debe lanzar TypeError

    def test_non_core_part_qualifies(self):
        p = self._product_with_bom("940360")  # mueble: no es core part
        result = calculate_product_origin(p, self.treaty, user=None)
        self.assertEqual(result["status"], "QUALIFIES")
        self.assertNotIn("automotive_core", result["detail"])
        JSONRenderer().render(result)


class PriceHistoryTest(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="ACME", slug="acme")
        self.product = Product.objects.create(
            tenant=self.tenant, sku="MAT-1", description="x", kind="material",
            hs_code="720839", unit_cost=Decimal("11.2070"), currency="USD")

    def test_no_log_on_precision_only_difference(self):
        """11.2070 vs 11.207008 NO es un cambio real (4 decimales)."""
        before = {"unit_cost": Decimal("11.2070"), "currency": "USD"}
        logs = log_product_changes(
            product=self.product, before=before,
            after={"unit_cost": Decimal("11.207008"), "currency": "USD"},
            source="bulk")
        self.assertEqual(len(logs), 0)

    def test_logs_real_price_change(self):
        logs = log_product_changes(
            product=self.product,
            before={"unit_cost": Decimal("11.2070"), "currency": "USD"},
            after={"unit_cost": Decimal("12.5000"), "currency": "USD"}, source="manual")
        self.assertEqual(len(logs), 1)
        self.assertEqual(logs[0].kind, ProductChangeLog.Kind.PRICE)


class BulkImportTest(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="ACME", slug="acme")

    def test_sku_uppercased_and_sin_precio_reported(self):
        res = import_products(self.tenant, [
            {"sku": "tst-a", "descripcion": "A", "costo_unitario": ""},
            {"sku": "TST-B", "descripcion": "B", "costo_unitario": "5.50"},
        ], user=None)
        self.assertEqual(res["creados"], 2)
        self.assertIn("TST-A", res["creados_skus"])          # guardado en mayúsculas
        self.assertIn("TST-A", res["sin_precio_skus"])       # fila sin precio
        self.assertNotIn("TST-B", res["sin_precio_skus"])
        self.assertTrue(Product.objects.filter(tenant=self.tenant, sku="TST-A").exists())

    def test_case_insensitive_update_does_not_duplicate(self):
        import_products(self.tenant, [{"sku": "tst-a", "costo_unitario": ""}], user=None)
        res = import_products(self.tenant, [{"sku": "TST-A", "costo_unitario": "9.99"}], user=None)
        self.assertEqual(res["actualizados"], 1)
        self.assertEqual(Product.objects.filter(tenant=self.tenant, sku__iexact="TST-A").count(), 1)
        self.assertEqual(
            Product.objects.get(tenant=self.tenant, sku="TST-A").unit_cost, Decimal("9.9900"))


class SolicitationCertificateTest(TestCase):
    def setUp(self):
        User = get_user_model()
        self.tenant = Tenant.objects.create(name="ACME", slug="acme")
        self.treaty = Treaty.objects.create(code="TMEC", name="USMCA",
                                            member_countries=["MX", "US", "CA"])
        self.product = Product.objects.create(tenant=self.tenant, sku="MAT-1",
                                              description="x", kind="material", hs_code="720839")
        self.supplier = Party.objects.create(tenant=self.tenant, kind=Party.Kind.SUPPLIER,
                                             name="Prov MX", country="MX")
        SupplierProfile.objects.create(tenant=self.tenant, party=self.supplier,
                                       legal_name="Prov MX SA",
                                       signature_png="data:image/png;base64,AAAA")
        self.decl = SupplierDeclaration.objects.create(
            tenant=self.tenant, supplier=self.supplier, product=self.product, treaty=self.treaty,
            is_originating=True, country_of_origin="MX",
            valid_from=date(2026, 1, 1), valid_to=date(2026, 12, 31))
        self.sr = SolicitationRequest.objects.create(
            tenant=self.tenant, supplier=self.supplier, product=self.product, treaty=self.treaty,
            status="responded", declaration=self.decl,
            period_from=date(2026, 1, 1), period_to=date(2026, 12, 31))
        self.company_user = User.objects.create(username="empresa")
        Membership.objects.create(user=self.company_user, tenant=self.tenant,
                                  role=Membership.Role.ADMIN)
        self.supplier_user = User.objects.create(username="prov")
        Membership.objects.create(user=self.supplier_user, tenant=self.tenant,
                                  role=Membership.Role.SUPPLIER, party=self.supplier)
        self.f = APIRequestFactory()

    def _accept(self, sign_method):
        req = self.f.post(f"/api/solicitations/{self.sr.id}/accept/",
                          {"sign_method": sign_method}, format="json")
        force_authenticate(req, user=self.company_user)
        return SolicitationRequestViewSet.as_view({"post": "accept"})(req, pk=self.sr.id)

    def _sign(self, body=None):
        cert = self.sr.certificate
        req = self.f.post(f"/api/solicitation-certificates/{cert.id}/sign/", body or {}, format="json")
        force_authenticate(req, user=self.supplier_user)
        return SolicitationCertificateViewSet.as_view({"post": "sign"})(req, pk=cert.id), cert

    def test_accept_creates_certificate_with_required_method(self):
        resp = self._accept("png_qr")
        self.assertEqual(resp.status_code, 200)
        self.sr.refresh_from_db()
        self.assertEqual(self.sr.status, "accepted")
        cert = self.sr.certificate
        self.assertEqual(cert.sign_method, "png_qr")
        self.assertFalse(cert.signed)
        self.assertEqual(cert.data["product"]["sku"], "MAT-1")
        self.assertTrue(cert.data["origin"]["is_originating"])

    def test_sign_png_qr_uses_profile_signature_and_token(self):
        self._accept("png_qr")
        resp, cert = self._sign()
        self.assertEqual(resp.status_code, 200)
        cert.refresh_from_db()
        self.assertTrue(cert.signed)
        self.assertTrue(cert.signature_png.startswith("data:image"))
        self.assertTrue(cert.verify_token)  # QR token generado

    def test_sign_manual_requires_scanned_file(self):
        self._accept("manual")
        resp, _ = self._sign({})  # sin archivo
        self.assertEqual(resp.status_code, 400)
        resp2, cert = self._sign({"scanned_file": "data:image/png;base64,ZZZZ"})
        self.assertEqual(resp2.status_code, 200)
        cert.refresh_from_db()
        self.assertTrue(cert.signed)
        self.assertTrue(cert.scanned_file.startswith("data:"))

    def test_company_cannot_sign(self):
        self._accept("png")
        cert = self.sr.certificate
        req = self.f.post(f"/api/solicitation-certificates/{cert.id}/sign/", {}, format="json")
        force_authenticate(req, user=self.company_user)
        resp = SolicitationCertificateViewSet.as_view({"post": "sign"})(req, pk=cert.id)
        self.assertEqual(resp.status_code, 403)

    def test_other_supplier_cannot_see_certificate(self):
        self._accept("png")
        User = get_user_model()
        other = Party.objects.create(tenant=self.tenant, kind=Party.Kind.SUPPLIER,
                                     name="Otro", country="MX")
        other_user = User.objects.create(username="otro")
        Membership.objects.create(user=other_user, tenant=self.tenant,
                                  role=Membership.Role.SUPPLIER, party=other)
        req = self.f.get("/api/solicitation-certificates/")
        force_authenticate(req, user=other_user)
        resp = SolicitationCertificateViewSet.as_view({"get": "list"})(req)
        results = resp.data.get("results", resp.data)
        self.assertEqual(len(results), 0)


class ClientLayoutTest(TestCase):
    def setUp(self):
        User = get_user_model()
        self.tenant = Tenant.objects.create(name="ACME", slug="acme")
        self.treaty = Treaty.objects.create(code="TMEC", name="USMCA",
                                            member_countries=["MX", "US", "CA"])
        self.client_party = Party.objects.create(tenant=self.tenant, kind="customer",
                                                 name="STELLANTIS", country="US")
        self.user = User.objects.create(username="emp")
        Membership.objects.create(user=self.user, tenant=self.tenant,
                                  role=Membership.Role.ADMIN)
        self.f = APIRequestFactory()

    def _template_bytes(self):
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Origin"
        ws.append(["Part Number", "Description", "HTS", "Qualified Y/N", "Criterion"])
        buf = BytesIO(); wb.save(buf); return buf.getvalue()

    def _upload(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from apps.origin.views import ClientLayoutViewSet
        up = SimpleUploadedFile("stellantis.xlsx", self._template_bytes())
        req = self.f.post("/api/client-layouts/", {
            "file": up, "client": self.client_party.id, "treaty": self.treaty.id,
            "name": "Portal STELLANTIS"}, format="multipart")
        force_authenticate(req, user=self.user)
        return ClientLayoutViewSet.as_view({"post": "create"})(req)

    def test_upload_detects_headers(self):
        resp = self._upload()
        self.assertIn(resp.status_code, (200, 201))
        self.assertEqual(resp.data["headers"]["A"], "Part Number")
        self.assertEqual(resp.data["headers"]["D"], "Qualified Y/N")
        self.assertEqual(resp.data["header_row"], 1)

    def test_mapping_validation_and_generate(self):
        from apps.origin.views import ClientLayoutViewSet
        from apps.origin.models import ClientOriginLayout, Qualification
        import openpyxl
        from io import BytesIO
        self._upload()
        layout = ClientOriginLayout.objects.get(tenant=self.tenant)
        # mapeo inválido rechazado
        req = self.f.patch(f"/api/client-layouts/{layout.id}/",
                           {"mapping": {"A": "no_existe"}}, format="json")
        force_authenticate(req, user=self.user)
        resp = ClientLayoutViewSet.as_view({"patch": "partial_update"})(req, pk=layout.id)
        self.assertEqual(resp.status_code, 400)
        # mapeo válido
        req = self.f.patch(f"/api/client-layouts/{layout.id}/", {"mapping": {
            "A": "sku", "B": "description", "C": "hs_formatted",
            "D": "origin_yn", "E": "criterion"}}, format="json")
        force_authenticate(req, user=self.user)
        resp = ClientLayoutViewSet.as_view({"patch": "partial_update"})(req, pk=layout.id)
        self.assertEqual(resp.status_code, 200)
        # producto calificado + generar
        prod = Product.objects.create(tenant=self.tenant, sku="P-1", description="Suspensión",
                                      kind="finished", hs_code="940360")
        Qualification.objects.create(tenant=self.tenant, product=prod, treaty=self.treaty,
                                     status="QUALIFIES", criterion="CTC")
        req = self.f.post(f"/api/client-layouts/{layout.id}/generate/",
                          {"products": [prod.id], "period_from": "2026-01-01",
                           "period_to": "2026-12-31"}, format="json")
        force_authenticate(req, user=self.user)
        resp = ClientLayoutViewSet.as_view({"post": "generate"})(req, pk=layout.id)
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb["Origin"]
        self.assertEqual(ws["A2"].value, "P-1")
        self.assertEqual(ws["C2"].value, "9403.60")
        self.assertEqual(ws["D2"].value, "Y")
        self.assertEqual(ws["E2"].value, "CTC")
