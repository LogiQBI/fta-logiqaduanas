"""Layouts del portal de origen del CLIENTE (ej. STELLANTIS).

La empresa sube la plantilla .xlsx de su cliente, mapea cada columna a un campo
de FTA y el sistema genera el archivo lleno (una fila por número de parte) con
los resultados de sus cálculos de origen, listo para subir al portal del cliente.
"""
import base64
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter

# Campos de FTA que se pueden mapear a una columna del layout del cliente.
# (clave, etiqueta para el desplegable del mapeo)
LAYOUT_FIELDS = [
    ("sku", "Núm. de parte (SKU)"),
    ("description", "Descripción"),
    ("hs", "Fracción HS (solo dígitos)"),
    ("hs_formatted", "Fracción HS (con punto, ej. 8708.80)"),
    ("origin_yn", "¿Originario? (Y/N)"),
    ("origin_sino", "¿Originario? (SÍ/NO)"),
    ("status", "Resultado de origen (texto)"),
    ("criterion", "Criterio de origen"),
    ("rvc", "VCR (%)"),
    ("country", "País de origen del producto"),
    ("treaty", "Tratado"),
    ("period_from", "Vigencia desde"),
    ("period_to", "Vigencia hasta"),
    ("date", "Fecha de generación"),
    ("company_name", "Razón social de tu empresa"),
    ("company_tax_id", "RFC / Tax ID de tu empresa"),
    ("company_country", "País de tu empresa"),
    ("client_name", "Nombre del cliente"),
    ("supplier_name", "Proveedor del producto"),
    ("unit_cost", "Costo unitario"),
    ("currency", "Moneda"),
]
LAYOUT_FIELD_KEYS = {k for k, _ in LAYOUT_FIELDS}

_TREATY_LABELS = {"TMEC": "USMCA"}


def read_template(file_bytes):
    """Lee la plantilla subida y detecta hoja + fila de encabezados (la primera
    con ≥2 celdas con texto). Devuelve (sheet_name, header_row, {col: encabezado})."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        cells = {get_column_letter(i + 1): str(v).strip()
                 for i, v in enumerate(row or []) if v not in (None, "")}
        if len(cells) >= 2:
            return ws.title, r_idx, cells
    return ws.title, 1, {}


def _fmt_hs(code):
    d = "".join(c for c in (code or "") if c.isdigit())
    return f"{d[:4]}.{d[4:6]}" if len(d) >= 6 else d


def _field_value(key, product, qual, ctx):
    """Valor de un campo de FTA para una fila (producto + su calificación)."""
    qualifies = bool(qual and qual.status == "QUALIFIES")
    if key == "sku":
        return product.sku
    if key == "description":
        return product.description
    if key == "hs":
        return "".join(c for c in (product.hs_code or "") if c.isdigit())
    if key == "hs_formatted":
        return _fmt_hs(product.hs_code)
    if key == "origin_yn":
        return "Y" if qualifies else "N"
    if key == "origin_sino":
        return "SÍ" if qualifies else "NO"
    if key == "status":
        return qual.get_status_display() if qual else "Sin calcular"
    if key == "criterion":
        return (qual.criterion or "") if qual else ""
    if key == "rvc":
        return str(qual.rvc_value) if (qual and qual.rvc_value is not None) else ""
    if key == "country":
        return product.country_of_origin or ""
    if key == "treaty":
        return ctx.get("treaty_label", "")
    if key in ("period_from", "period_to", "date"):
        return ctx.get(key, "")
    if key in ("company_name", "company_tax_id", "company_country", "client_name"):
        return ctx.get(key, "")
    if key == "supplier_name":
        return product.supplier.name if product.supplier_id else ""
    if key == "unit_cost":
        return str(product.unit_cost or "")
    if key == "currency":
        return product.currency or ""
    return ""


def generate_layout(layout, rows, ctx):
    """Llena la plantilla del cliente con una fila por producto y devuelve los
    bytes del .xlsx resultante.

    `rows`: lista de (product, qualification|None).
    `ctx`: valores comunes {treaty_label, period_from, period_to, date,
    company_name, company_tax_id, company_country, client_name}."""
    file_bytes = base64.b64decode(layout.file_b64 or "")
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb[layout.sheet_name] if layout.sheet_name in wb.sheetnames else wb.active
    mapping = {col: key for col, key in (layout.mapping or {}).items()
               if key in LAYOUT_FIELD_KEYS}
    start = (layout.header_row or 1) + 1
    for i, (product, qual) in enumerate(rows):
        for col, key in mapping.items():
            ws[f"{col}{start + i}"] = _field_value(key, product, qual, ctx)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
